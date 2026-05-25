"""Generate demo Excel workbook for propulsion budget."""

from pathlib import Path

from openpyxl import Workbook

OUT = Path(__file__).parent / "propulsion_budget.xlsx"


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["A2"] = "DC bus voltage (V)"
    ws["B2"] = 400
    ws["A3"] = "Motor mass (kg)"
    ws["B3"] = 12.4
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
