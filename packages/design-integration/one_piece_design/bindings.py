"""Excel cell ↔ Python marker bindings — minimum SSOT integration unit."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

MARKER_PATTERN = re.compile(
    r"^(?P<prefix>#?\s*)SSOT:(?P<kind>PARAM|REQ):(?P<key>[A-Z0-9_-]+)\s*=\s*(?P<value>.+)$",
    re.MULTILINE,
)


@dataclass(frozen=True)
class BindingSpec:
    excel_cell_ref: str
    python_marker: str
    design_parameter_key: str
    """Executable variable name synced alongside the SSOT marker comment."""
    python_assignment_name: str | None = None


def read_excel_cell(workbook_path: Path, cell_ref: str) -> str | float | bool:
    """Read a single cell value from an Excel workbook (e.g. 'Inputs!B4')."""
    if "!" in cell_ref:
        sheet_name, coord = cell_ref.split("!", 1)
    else:
        sheet_name, coord = None, cell_ref

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    value = ws[coord].value
    if value is None:
        return ""
    return value


def parse_python_markers(script_path: Path) -> dict[str, str]:
    """Return marker key → current literal value in the Python script."""
    text = script_path.read_text(encoding="utf-8")
    result: dict[str, str] = {}
    for match in MARKER_PATTERN.finditer(text):
        key = match.group("key")
        result[key] = match.group("value").strip()
    return result


def _format_literal(raw: str | float | bool | int) -> str:
    if isinstance(raw, bool):
        return "True" if raw else "False"
    if isinstance(raw, (int, float)):
        if isinstance(raw, float) and raw != int(raw):
            return repr(raw)
        return str(int(raw))
    return repr(str(raw))


def _literals_equivalent(old: str, new: str) -> bool:
    """Best-effort compare of Python literal strings."""
    if old == new:
        return True
    try:
        return eval(old, {"__builtins__": {}}, {}) == eval(new, {"__builtins__": {}}, {})  # noqa: S307
    except Exception:
        return False


def _marker_pattern(marker_key: str) -> re.Pattern[str]:
    return re.compile(
        rf"^(\s*#\s*SSOT:PARAM:{re.escape(marker_key)}\s*=\s*).+$",
        re.MULTILINE,
    )


def _assignment_pattern(var_name: str) -> re.Pattern[str]:
    return re.compile(
        rf"^(\s*{re.escape(var_name)}\s*=\s*).+$",
        re.MULTILINE,
    )


def _literal_from_match(match: re.Match[str]) -> str:
    return match.group(0).split("=", 1)[1].strip()


def _binding_targets_in_sync(
    text: str,
    marker_key: str,
    new_literal: str,
    assignment_name: str | None,
) -> bool:
    marker_match = _marker_pattern(marker_key).search(text)
    if marker_match is None:
        return False
    if not _literals_equivalent(_literal_from_match(marker_match), new_literal):
        return False
    if assignment_name is None:
        return True
    assignment_match = _assignment_pattern(assignment_name).search(text)
    if assignment_match is None:
        return False
    return _literals_equivalent(_literal_from_match(assignment_match), new_literal)


def sync_python_from_excel(
    workbook_path: Path,
    script_path: Path,
    bindings: list[BindingSpec],
) -> list[str]:
    """
    Propagate Excel cell values into Python SSOT markers and executable assignments.
    Returns list of human-readable change summaries.
    """
    text = script_path.read_text(encoding="utf-8")
    changes: list[str] = []

    for binding in bindings:
        raw = read_excel_cell(workbook_path, binding.excel_cell_ref)
        new_literal = _format_literal(raw)
        marker_key = binding.design_parameter_key

        if _binding_targets_in_sync(
            text,
            marker_key,
            new_literal,
            binding.python_assignment_name,
        ):
            continue

        marker = _marker_pattern(marker_key)
        marker_match = marker.search(text)
        if marker_match is None:
            raise ValueError(
                f"Marker SSOT:PARAM:{marker_key} not found in {script_path}",
            )

        old_marker = _literal_from_match(marker_match)

        def marker_replacer(m: re.Match[str]) -> str:
            return f"{m.group(1)}{new_literal}"

        text = marker.sub(marker_replacer, text, count=1)
        summary = (
            f"{binding.excel_cell_ref} → SSOT:PARAM:{marker_key}: "
            f"{old_marker} → {new_literal}"
        )

        if binding.python_assignment_name:
            assignment = _assignment_pattern(binding.python_assignment_name)
            assignment_match = assignment.search(text)
            if assignment_match is None:
                raise ValueError(
                    f"Assignment {binding.python_assignment_name} not found in {script_path}",
                )
            old_assignment = _literal_from_match(assignment_match)

            def assignment_replacer(m: re.Match[str]) -> str:
                return f"{m.group(1)}{new_literal}"

            text = assignment.sub(assignment_replacer, text, count=1)
            summary += (
                f"; {binding.python_assignment_name}: "
                f"{old_assignment} → {new_literal}"
            )

        changes.append(summary)

    if changes:
        script_path.write_text(text, encoding="utf-8")

    return changes
