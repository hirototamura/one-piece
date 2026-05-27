"""Tests for Excel ↔ Python SSOT bindings."""

from pathlib import Path

import pytest

from one_piece_design.bindings import (
    BindingSpec,
    ValueBindingSpec,
    sync_python_from_excel,
    sync_python_from_values,
)
from one_piece_design.policy import can_actor_mutate
from one_piece_design.runner import run_python_script

EXAMPLES = Path(__file__).resolve().parent.parent / "examples"


@pytest.fixture(scope="module")
def workbook(tmp_path_factory):
    tmp = tmp_path_factory.mktemp("wb")
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["B2"] = 400
    ws["B3"] = 12.4
    path = tmp / "budget.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def script_copy(tmp_path):
    src = EXAMPLES / "thrust_margin.py"
    dest = tmp_path / "thrust_margin.py"
    dest.write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
    return dest


@pytest.fixture
def thermal_script_copy(tmp_path):
    src = EXAMPLES / "thermal_rejection.py"
    dest = tmp_path / "thermal_rejection.py"
    dest.write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
    return dest


def test_can_actor_mutate_defaults():
    assert can_actor_mutate("human_engineer", "requirement", "critical")
    assert not can_actor_mutate("ai_agent", "requirement", "critical")
    assert can_actor_mutate("ai_agent", "design_parameter", "standard")
    assert not can_actor_mutate(
        "ai_agent",
        "design_parameter",
        "standard",
        ai_allowed_fraction=0,
    )
    assert can_actor_mutate("logic_automation", "design_parameter", "derived")


def test_sync_python_from_excel(workbook, script_copy):
    bindings = [
        BindingSpec("Inputs!B2", "SSOT:PARAM:P-VBUS", "P-VBUS", "VBUS"),
        BindingSpec("Inputs!B3", "SSOT:PARAM:P-M-MOTOR", "P-M-MOTOR", "MOTOR_MASS_KG"),
    ]
    changes = sync_python_from_excel(workbook, script_copy, bindings)
    assert changes == []  # values already match

    from openpyxl import load_workbook

    wb = load_workbook(workbook)
    wb.active["B2"] = 410
    wb.save(workbook)

    changes = sync_python_from_excel(workbook, script_copy, bindings)
    assert any("410" in c for c in changes)
    text = script_copy.read_text(encoding="utf-8")
    assert "SSOT:PARAM:P-VBUS = 410" in text
    assert "VBUS = 410" in text

    result = run_python_script(script_copy)
    assert result.success
    assert "nominal bus 410 V" in result.stdout


def test_run_python_script(script_copy):
    result = run_python_script(script_copy)
    assert result.success
    assert "Thrust margin check" in result.stdout


def test_sync_python_from_values(thermal_script_copy):
    changes = sync_python_from_values(
        thermal_script_copy,
        [
            ValueBindingSpec("P-RAD-AREA", 5.9, "RADIATOR_AREA_M2"),
            ValueBindingSpec("P-RAD-MASS", 15.9, "RADIATOR_MASS_KG"),
            ValueBindingSpec("P-COOLANT-FLOW", 0.87, "COOLANT_FLOW_KGPS"),
        ],
    )
    assert len(changes) == 3

    text = thermal_script_copy.read_text(encoding="utf-8")
    assert "SSOT:PARAM:P-RAD-AREA = 5.9" in text
    assert "RADIATOR_MASS_KG = 15.9" in text

    result = run_python_script(thermal_script_copy)
    assert result.success
    assert "Thermal rejection check" in result.stdout
    assert "METRICS:" in result.stdout
